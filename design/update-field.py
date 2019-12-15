# -*- coding: utf-8 -*-
#!/usr/bin/python

import argparse
import csv
import copy
import logging
import sys
import shutil
import os
import openpyxl as pyxl

from kifield import sch
from kifield import kifield

logger = logging.getLogger('kifield')

def fill_quantity_ws(ws):
    ws.cell(row=1, column=1).value = "refs"
    ws.cell(row=1, column=2).value = "footprint"
    ws.cell(row=1, column=3).value = "value"
    ws.cell(row=1, column=4).value = "spn1"
    ws.cell(row=1, column=5).value = "supplier1"
    ws.cell(row=1, column=6).value = "quantity"
    ws.cell(row=1, column=7).value = 'price1'
    ws.cell(row=1, column=8).value = 'price100'
    ws.cell(row=1, column=9).value = 'price1000'

def make_authority_ref(fields):
    i = len(fields)
    while True:
        new_ref = "F" + str(i)
        if new_ref not in fields:
            return new_ref
        i += 1

def get_part_key(part):
    return part['footprint'] + '~' + part['value']

def get_all_part_keys(fields):
    keys = []
    for ref in fields:
        key = get_part_key(fields[ref])
        if key not in keys:
            keys.append(key)
    keys.sort()
    return keys

def get_refs_by_key(fields):
    refs_by_key = {}
    for ref in fields:
        key = get_part_key(fields[ref])
        refs_by_key[key] = refs_by_key.get(key, []) + [ref]
    return refs_by_key

def update_fields(filename, authority_fields, quantities, uses, ws, add_missing=False):
    all_child_fields = kifield.extract_part_fields_from_sch(filename, recurse=True)
    authority_fields_by_key = {}
    for ref in authority_fields:
        authority_fields_by_key[get_part_key(authority_fields[ref])] = authority_fields[ref]

    name = os.path.basename(filename)
    modified = False
    for row, ref in enumerate(all_child_fields):
        child_fields = all_child_fields[ref]
        key = get_part_key(child_fields)
        quantities[key] = quantities.get(key, 0) + 1
        if name not in uses.get(key, []):
            uses[key] = uses.get(key, []) + [name]
        old_spn = child_fields.get('spn1', "")
        fp = child_fields['footprint']
        value = child_fields['value']
        if key in authority_fields_by_key:
            new_spn = authority_fields_by_key[key].get('spn1', "")
            if new_spn != old_spn:
                modified = True
                logger.log(logging.INFO, "Fixing %s (%s vs %s)" % (ref, old_spn, new_spn))
                child_fields['spn1'] = new_spn
            pass
        else:
            if not add_missing:
                raise Exception("No authority on part: %s %s (%s)" % (ref, fp, old_spn))
            else:
                new_ref = make_authority_ref(authority_fields)
                logger.log(logging.INFO, "Adding missing part authority: %s value=%s fp=%s spn1=%s (%s)" % (ref, value, fp, old_spn, new_ref))
                authority_fields[new_ref] = child_fields
                authority_fields_by_key[key] = child_fields

    if modified:
        logger.log(logging.INFO, "Saving %s" % (filename))
        backup_file(filename)
        kifield.insert_part_fields_into_sch(all_child_fields, filename, True, False, False)

    fill_quantity_ws(ws)
    refs_by_key = get_refs_by_key(all_child_fields)
    all_keys = get_all_part_keys(all_child_fields)
    for row, key in enumerate(all_keys):
        all_refs = refs_by_key[key]
        ws.cell(row=row + 2, column=1).value = ','.join(all_refs)
        ws.cell(row=row + 2, column=2).value = all_child_fields[all_refs[0]]['footprint']
        ws.cell(row=row + 2, column=3).value = all_child_fields[all_refs[0]]['value']
        ws.cell(row=row + 2, column=4).value = all_child_fields[all_refs[0]].get('spn1')
        ws.cell(row=row + 2, column=5).value = all_child_fields[all_refs[0]].get('supplier1')
        ws.cell(row=row + 2, column=6).value = len(all_refs)
        ws.cell(row=row + 2, column=7).value = authority_fields_by_key[key].get('price1', 0.0)
        ws.cell(row=row + 2, column=8).value = authority_fields_by_key[key].get('price100', 0.0)
        ws.cell(row=row + 2, column=9).value = authority_fields_by_key[key].get('price1000', 0.0)

    return authority_fields


def update_field(filename, field, oldValue, newValue, recurse=False):
    all_child_fields = kifield.extract_part_fields_from_sch(filename, recurse=recurse)
    for row, ref in enumerate(all_child_fields):
        child_fields = all_child_fields[ref]
        if child_fields[field] == oldValue:
            child_fields[field] = newValue

    kifield.insert_part_fields_into_sch(all_child_fields, filename, recurse, False, False)

def backup_file(filename):
    index = 1
    while True:
        backup_filename = filename + '.{}.bak'.format(index, file)
        if not os.path.isfile(backup_filename):
            shutil.copy(filename, backup_filename)
            break
        index += 1

def write_authority(filename, fields):
    authority_fields = ['footprint', 'value', 'spn1', 'supplier1', 'spn2', 'supplier2', 'price1', 'price100', 'price1000']
    with open(filename + ".csv", 'wb') as f:
        logger.log(logging.INFO, "Writing %s..." % (filename))
        w = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        w.writerow(['refs'] + authority_fields)
        refs_by_key = get_refs_by_key(fields)
        all_keys = get_all_part_keys(fields)
        for row, key in enumerate(all_keys):
            ref = refs_by_key[key][0]
            values = [ 'F' + str(row) ]
            part_fields = fields[ref]
            for name in authority_fields:
                if name in part_fields:
                    values.append(part_fields[name])
                else:
                    values.append('')
            w.writerow(values)

def check_authority(authority_fields):
    seen_keys = {}
    for ref in authority_fields:
        key = get_part_key(authority_fields[ref])
        if key in seen_keys:
            print(key)

def main():
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)

    # update_field("VESC_6.sch", 'footprint', 'CRF1:SMD-0603_r', 'Resistors_SMD:R_0603', True)
    # update_field("VESC_6.sch", 'footprint', 'CRF1:SMD-0603_c', 'Capacitors_SMD:C_0603', True)
    # update_field("VESC_6.sch", 'footprint', 'w_smd_cap:c_1210', 'Capacitors_SMD:C_1210', True)

    # update_field("VESC_6.sch", 'footprint', 'Resistor_SMD:R_0603_1608Metric', 'Resistors_SMD:R_0603', True)
    # update_field("VESC_6.sch", 'footprint', 'Capacitor_SMD:C_0603_1608Metric', 'Capacitors_SMD:C_0603', True)

    update_field("VESC_6.sch", 'footprint', 'Resistors_SMD:R_0603', 'Resistor_SMD:R_0603_1608Metric', True)
    update_field("VESC_6.sch", 'footprint', 'Capacitors_SMD:C_0603', 'Capacitor_SMD:C_0603_1608Metric', True)

    update_field("VESC_6.sch", 'footprint', 'LEDs:LED_0603', 'LED_SMD:LED_0603_1608Metric', True)


if __name__ == "__main__":
    main()
